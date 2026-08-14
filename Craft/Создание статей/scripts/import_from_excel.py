# -*- coding: utf-8 -*-
"""Импорт статей с примерами вопросов из Excel в базу знаний Craft-Talk.

Порядок для каждой темы (см. docs/):
    1. создать/обновить тему (article/update, корневая, Answers: []);
    2. опубликовать тему (publish), т.к. вложенные создаются только под Active;
    3. взять категорию темы из дерева (/catalog/categories, root) — её Id/SymbolCode;
    4. для каждой строки: создать/обновить статью-вопрос (answers), добавить разметку
       (/markup/add — примеры вопросов = разметка интентов).

Использование:
    $env:CRAFTTALK_TOKEN='<token>'
    python import_from_excel.py config_21vek.json
    python import_from_excel.py config_21vek.json --limit 2        # только первые 2 строки
    python import_from_excel.py config_21vek.json --theme "Бонусные баллы"

Токен — из env CRAFTTALK_TOKEN, в конфиг/код не класть.
"""
import csv
import json
import sys
import time
import uuid

import openpyxl

from crafttalk_client import (
    HOST, api, norm, article_payload, parse_article, find_category_by_title,
    find_article_by_ext,
)


def load_config(path):
    with open(path, encoding="utf-8") as f:
        cfg = json.load(f)
    return cfg


def load_rows(cfg):
    """Читает Excel: возвращает список dict {row, theme, question, answer, examples[]}."""
    wb = openpyxl.load_workbook(cfg["xlsx"], data_only=True)
    ws = wb[cfg["sheet"]]
    cols = cfg["columns"]
    rng = cfg["rows"]
    rows = []
    for r in range(rng["start"], rng["end"] + 1):
        if r in (rng.get("skip") or []):
            continue
        question = norm(ws.cell(r, cols["question"]).value)
        if not question:
            continue
        theme = norm(ws.cell(r, cols["theme"]).value)
        answer = norm(ws.cell(r, cols["answer"]).value)
        examples = [e.strip() for e in str(ws.cell(r, cols["examples"]).value or "").splitlines() if e.strip()]
        rows.append({"row": r, "theme": theme, "question": question, "answer": answer, "examples": examples})
    return rows


def ordered_themes(rows):
    seen, out = set(), []
    for r in rows:
        if r["theme"] and r["theme"] not in seen:
            seen.add(r["theme"])
            out.append(r["theme"])
    return out


def make_ext(cfg, kind, n=None, row=None):
    pat = cfg["ext_ids"][kind].format(src=cfg["ext_source"], n=n, row=row)
    return pat


def run(cfg, only_rows=None, only_theme=None, add_markup=True, publish_themes=True, sleep=0.3):
    import os
    if not os.environ.get("CRAFTTALK_TOKEN"):
        print("ERROR: CRAFTTALK_TOKEN env var is not set")
        sys.exit(1)

    project, catalog, src = cfg["project"], cfg["catalog_code"], cfg["ext_source"]
    rows = load_rows(cfg)
    if only_rows:
        rows = [r for r in rows if r["row"] in only_rows]
    if only_theme:
        rows = [r for r in rows if r["theme"] == only_theme]
    if not rows:
        print("no rows to process")
        return
    themes = ordered_themes(load_rows(cfg))  # глобальный порядок тем (для стабильных ExtId)
    themes_to_do = ordered_themes(rows)

    report = []
    for idx, theme in enumerate(themes, start=1):
        if theme not in themes_to_do:
            continue
        text_id = make_ext(cfg, "theme", n=idx)
        payload = article_payload(project, catalog, text_id, src, theme,
                                  catalog, "root", has_children=True)
        status, body = api("POST", "/article/update", payload)
        aid, sc = parse_article(body)

        parent_id, parent_code = aid, sc
        if publish_themes and aid:
            pst, pbody = api("POST", "/article/publish",
                             {"ProjectId": project, "ExtId": text_id, "ExtSourceId": src})
            time.sleep(sleep)
            cat = find_category_by_title(project, catalog, theme, status="Active")
            if cat:
                parent_id, parent_code = cat.get("Id"), cat.get("SymbolCode")

        msg = ""
        if status != 200:
            msg = str(body)[:300]
        elif not parent_id:
            msg = "parent resolution failed"
        print("THEME  %-2d %s [%s] parent=%s %s" % (idx, theme, status, parent_id, msg))
        report.append(["", theme, theme, "theme", text_id, status, aid, sc, "", "", "", msg])
        time.sleep(sleep)
        if not parent_id:
            continue

        for r in [x for x in rows if x["theme"] == theme]:
            qext = make_ext(cfg, "question", row=r["row"])
            answers = []
            if r["answer"]:
                answers = [{"Id": str(uuid.uuid5(uuid.NAMESPACE_DNS, qext)),
                            "Text": r["answer"], "Slots": []}]
            payload = article_payload(project, catalog, qext, src, r["question"],
                                      parent_id, parent_code, "root", False, answers)
            qstatus, qbody = api("POST", "/article/update", payload)
            qid, qsc = parse_article(qbody)

            if qstatus == 400:  # Removed-статья: обновляем по Id
                art = find_article_by_ext(project, qext, src)
                if art and art.get("Id"):
                    payload = article_payload(project, catalog, qext, src, r["question"],
                                              parent_id, parent_code, "root", False,
                                              answers, article_id=art["Id"])
                    qstatus, qbody = api("POST", "/article/update", payload)
                    qid, qsc = parse_article(qbody)

            qmsg, batch_id = "", ""
            if qstatus == 200 and qsc and r["examples"] and add_markup:
                markups = [{"MarkupText": e, "IntentSymbolCode": qsc} for e in r["examples"]]
                mst, mbody = api("POST", "/markup/add", {"ProjectId": project, "Markups": markups})
                if mst == 200 and isinstance(mbody, dict):
                    batch_id = mbody.get("MarkupBatchId", "") or ""
                else:
                    qmsg = "markup failed: %s" % str(mbody)[:150]
                time.sleep(sleep)
            elif qstatus != 200:
                qmsg = str(qbody)[:300]

            print("  Q %-4d [%s] %s %s" % (r["row"], qstatus, qid, qmsg))
            report.append([r["row"], theme, r["question"], "question", qext, qstatus,
                           qid, qsc, r["answer"][:60], str(len(r["examples"])), batch_id, qmsg])
            time.sleep(sleep)

    report_path = cfg.get("report") or "import_report.csv"
    with open(report_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Row", "Theme", "Title", "Type", "ExtId", "HTTP", "ArticleId",
                    "SymbolCode", "Answer", "Examples", "MarkupBatchId", "Error"])
        w.writerows(report)
    print("REPORT: %s" % report_path)
    print("host:", HOST)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    cfg_path = sys.argv[1]
    cfg = load_config(cfg_path)
    args = sys.argv[2:]
    limit = None
    theme_only = None
    if "--limit" in args:
        limit = int(args[args.index("--limit") + 1])
    if "--theme" in args:
        theme_only = args[args.index("--theme") + 1]
    only = None
    if limit is not None:
        all_rows = load_rows(cfg)
        only = {r["row"] for r in all_rows[:limit]}
    run(cfg, only_rows=only, only_theme=theme_only)
