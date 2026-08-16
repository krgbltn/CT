# -*- coding: utf-8 -*-
"""Сборка файла «Необходимые интеграции.xlsx»: 13 листов.

Листы интеграций (8 скриптов + идентификация): № | Лист внутри скрипта | Вопрос
| ID статьи | Что запросить | Подходящий метод | Метод API (ключ) | Нужные поля из ответа
| JSONPath для slotsMapping | ID слотов которые заполняем | ID агента | Статус | Описание агента.
Справочники: «Методы API» (со сценариями использования), «Слоты», «Агенты».
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = r"d:\сигурд\готовые скрипты\Необходимые интеграции.xlsx"

HEADER_SCRIPT = ["№", "Лист внутри скрипта", "Вопрос", "ID статьи",
                 "Что запросить в интеграции", "Подходящий метод",
                 "Метод API (ключ)", "Нужные поля из ответа",
                 "JSONPath для slotsMapping", "ID слотов которые заполняем",
                  "ID агента", "Статус", "Описание агента",
                  "Перевод на оператора при ошибке"]
HEADER_METHODS = ["№", "Метод", "Тип", "URL/Операция", "Входные параметры",
                  "Назначение", "Поля ответа", "Сценарии использования"]
HEADER_SLOTS = ["ID слота", "Тип", "Описание / возможные значения",
                "Сценарии, где слот ЗАПОЛНЯЕТСЯ", "Сценарии, где слот ИСПОЛЬЗУЕТСЯ"]
HEADER_AGENTS = ["ID агента", "Статус", "Метод API (ключ)",
                 "Вопросы, в которых вызывается", "Объединение вызовов",
                 "Перевод на оператора при ошибке"]

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_C = Alignment(wrap_text=True, vertical="top", horizontal="center")

FILL_REST = PatternFill("solid", fgColor="E2EFDA")     # зелёный — REST
FILL_SOAP = PatternFill("solid", fgColor="DDEBF7")     # синий — SOAP
FILL_ASK = PatternFill("solid", fgColor="FFF2CC")      # жёлтый — уточняется у заказчика
FILL_NONE = PatternFill("solid", fgColor="D9D9D9")     # серый — не требуется / вне API

# Единый фолбэк всех агентов: при неуспешном запросе — перевод на оператора
FALLBACK_OPERATOR = '/switchredirect aiassist2 intent_id="article-d6585ce7-c4e9-4e42-97d2-bc4142e8ae1d"'

WIDTHS_SCRIPT = [5, 22, 32, 8, 34, 26, 20, 30, 36, 18, 24, 20, 58, 40]
WIDTHS_METHODS = [5, 28, 10, 44, 30, 46, 46, 42]
WIDTHS_SLOTS = [18, 26, 52, 62, 44]
WIDTHS_AGENTS = [26, 24, 24, 84, 72, 44]


# --- JSONPath-пути для slotsMapping по агентам (slotId -> путь в ответе) ---
AGENT_SLOT_PATHS = {
    "identification": {
        "user_id": "[0].UserId",
        "contract_no": "[0].WebProperties.ContractNo",
        "address": "[0].WebProperties.Address",
        "phone": "входящий номер (message)",
        "house_type": "housetype (GET /disconnection_report_info)",
    },
    "get_client_info": {
        "user_id": "id",
        "contract_no": "account.number",
        "address": "account.house.address",
        "owner_fio": "account.owner",
        "status": "account.status",
        "balance": "account.balance",
        "is_owner_registered": "account.is_owner_registered",
        "service": "account.accounts[]",
        "device_status": "account.devices[]",
        "email": "email",
    },
    "contracts_by_number": {
        "user_id": "[0].UserId",
        "contract_no": "[0].WebProperties.ContractNo",
        "address": "[0].WebProperties.Address",
    },
    "contracts_by_address_lastname": {
        "user_id": "[0].UserId",
        "contract_no": "[0].WebProperties.ContractNo",
    },
    "contracts_by_address": {
        "user_id": "[0].UserId",
        "contract_no": "[0].WebProperties.ContractNo",
        "address": "[0].WebProperties.Address",
    },
    "contracts_by_number_lastname": {
        "user_id": "[0].UserId",
        "contract_no": "[0].WebProperties.ContractNo",
    },
    "soap_contracts_by_phone": {
        "user_id": "ContractInfo.ID",
        "contract_no": "ContractInfo.No",
        "address": "ContractInfo.Adress",
        "owner_fio": "ContractInfo.AbonentName",
    },
    "disabling_electro": {
        "disconnection": "disconnection.text",
        "debt": "debt",
        "house_type": "house_type",
    },
    "disabling_tsn": {
        "disconnection": "disconnection.text",
        "debt": "debt",
        "house_type": "house_type",
        "notifications": "notifications[] (GET /notifications)",
    },
    "disabling_tsn_plan": {
        "disconnection": "disconnection.text",
        "debt": "debt",
    },
    "get_balance": {
        "balance": "account.balance",
    },
    "get_transactions": {
        "last_payment": "transactions[0].transactions[0] (payment; date)",
    },
    "network_org": {
        "network_org": "Name (контакты — Contacts[])",
    },
    "devices_check": {
        "device_status": "[0].status",
        "last_reading": "[0].last_reading",
        "next_check": "[0].next_check",
        "is_smart": "[0].is_smart",
        "service": "[0].service_name",
    },
    "specialist_requests_check": {
        "request_status": "[0] (id; request_date; title)",
    },
    "device_history_get": {
        "readings_history": "[0].value (+ date; status_transfer)",
        "last_reading": "[0].value",
    },
    "device_verification_check": {
        "next_check": "[0].next_check / MDWarningInfo.NextVerificationDeadline",
    },
    "get_client_email": {
        "email": "email (поля нет в текущей модели /info — уточняется)",
    },
    "legal_contracts_by_inn": {
        "contract_no": "ContractInfo.No",
        "user_id": "ContractInfo.ID",
    },
    "legal_balance": {
        "balance": "ContractAddressAndBalanceByDicServiceGroupAndContragent.Balance",
    },
}

# --- Входные слоты по агентам (что агент читает из slot_context) ---
AGENT_INPUT_SLOTS = {
    "identification": ["phone"],
    "get_client_info": ["user_id"],
    "disabling_electro": ["user_id"],
    "disabling_tsn": ["user_id"],
    "disabling_tsn_plan": ["user_id"],
    "get_balance": ["user_id"],
    "get_transactions": ["user_id"],
    "ksrt_report_create": ["user_id"],
    "ksrt_cancel": ["user_id"],
    "network_org": ["user_id"],
    "devices_check": ["user_id"],
    "specialist_requests_check": ["user_id"],
    "device_history_get": ["user_id"],
    "device_verification_check": ["user_id"],
    "reconciliation_act": ["user_id"],
    "soap_contracts_by_phone": ["phone"],
    "legal_balance": ["phone"],
    "legal_contracts_by_inn": ["phone"],
}

# --- Словарь слотов: ID -> (тип, описание / возможные значения) ---
SLOTS_INFO = {
    "user_id": ("string (UUID)", "Идентификатор лицевого счёта (UserId). Ключевой вход всех запросов /{user_id}/..."),
    "contract_no": ("string", "Номер лицевого счёта/договора (напр. ЕТСОО167639)"),
    "address": ("string", "Адрес ЛС (полный, с индексом). Используется для подтверждения адреса клиентом"),
    "phone": ("string", "Номер телефона клиента (из входящего вызова IVR или со слов клиента)"),
    "owner_fio": ("string", "ФИО владельца/абонента ЛС (для проверки «на кого оформлен ЛС» и писем)"),
    "status": ("string", "Статус ЛС/пользователя (Действует, Закрыт и т.д.)"),
    "balance": ("number", "Баланс/задолженность по ЛС (account.balance)"),
    "is_owner_registered": ("boolean", "Признак регистрации собственника в ЛК/МП (account.is_owner_registered)"),
    "service": ("string[]", "Услуги на расчётах (account.accounts[] или devices[].service_name)"),
    "device_status": ("string", "Статус прибора учёта (на расчётах / снят / принимает показания)"),
    "disconnection": ("string", "Текст отключения (disconnection.text) — для озвучивания клиенту и письма на network@"),
    "debt": ("boolean", "Наличие ограничения/наряда в связи с задолженностью"),
    "house_type": ("enum (OTHER/MKD/PRIVATE)", "Тип здания — для ветвлений сценария (частный дом / МКД / другое)"),
    "last_payment": ("object", "Последняя оплата: сумма, дата, способ поступления (transactions[0].transactions[0])"),
    "network_org": ("string", "Наименование сетевой организации (Name) + контакты (Contacts[])"),
    "is_smart": ("boolean", "Признак интеллектуального прибора учёта (devices[].is_smart)"),
    "last_reading": ("number", "Последние переданные показания ПУ (devices[].last_reading / device_history)"),
    "next_check": ("string (дата)", "Дата следующей поверки ПУ / истечения МПИ (devices[].next_check)"),
    "readings_history": ("object[]", "История передачи показаний: value, date, status_transfer"),
    "request_status": ("object", "Заявка на специалиста: id, request_date, title, статус (specialist_requests)"),
    "notifications": ("object[]", "Уведомления/оповещения по ЛС (GET /notifications) — для проверки оповещения"),
    "contract_open_date": ("string (дата)", "Дата открытия ЛС — поле/метод уточняется у заказчика"),
    "receipt_paper_flag": ("boolean", "Признак «Выставлять бумажную квитанцию» — метод уточняется у заказчика"),
    "tp_status": ("string", "Статус договора техприсоединения (ЛК ИЭСК) — метод уточняется у заказчика"),
    "email": ("string", "E-mail клиента — поля нет в текущей модели /info, метод уточняется у заказчика"),
    "residents": ("string[]", "Список жильцов/зарегистрированных по адресу — метод уточняется у заказчика"),
    "inn": ("string", "ИНН юридического лица — поиск по ИНН уточняется у заказчика"),
    "accountant": ("string", "Ответственный расчётчик по договору (АСУСЭиРП) — уточняется у заказчика"),
    "service_zone": ("string", "Зона обслуживания сетевой организации по адресу — уточняется у заказчика"),
    "final_answer": ("enum (1/2/3/error)", "Признак результата для статьи: 1 — отключение/наряд, 2 — долг, 3 — нет ничего, error — ошибка/нет данных. Заполняют все агенты"),
    "id_step": ("enum", "Внутренний слот машины состояний агента identification (confirm_single / ask_multiple)"),
    "id_selected_index": ("string", "Внутренний слот identification: индекс выбранного контракта"),
    "id_contracts_data": ("string (JSON)", "Внутренний слот identification: сериализованный список контрактов"),
}


def slot_list(s):
    return [x.strip() for x in (s or "").split(";") if x.strip() and x.strip() != "—"]


def scenario_label(sheet, section, question):
    if sheet == "Идентификация клиента":
        return "Идентификация клиента: %s" % (question or "")
    return "%s — %s: %s" % (sheet, section, question)


def derive_method_key(method):
    v = (method or "").lower()
    if v.startswith("метод уточняется") or v.startswith("уточнить"):
        return "—"
    if v == "не требуется" or v.startswith("вне api"):
        return "—"
    pairs = [
        ("contracts/by_number_lastname", "contracts_by_number_lastname"),
        ("contracts/by_address_lastname", "contracts_by_address_lastname"),
        ("contracts/by_number", "contracts_by_number"),
        ("contracts/by_phone", "contracts_by_phone"),
        ("contracts/by_address", "contracts_by_address"),
        ("contracts_by_number_lastname", "contracts_by_number_lastname"),
        ("contracts_by_address_lastname", "contracts_by_address_lastname"),
        ("contracts_by_number", "contracts_by_number"),
        ("contracts_by_phone", "contracts_by_phone"),
        ("contracts_by_address", "contracts_by_address"),
        ("devices/history", "device_history"),
        ("disconnections_electro", "disconnections_electro"),
        ("disconnections_heat", "disconnections_heat"),
        ("device_history", "device_history"),
        ("devices", "devices"),
        ("transactions", "transactions"),
        ("specialist_requests", "specialist_requests"),
        ("ksrt_requests", "ksrt_requests"),
        ("network_organization_data", "network_organization_data"),
        ("disconnection_report_info", "disconnection_report_info"),
        ("cancel_ksrt", "cancel_ksrt"),
        ("notifications_electro", "notifications_electro"),
        ("notifications", "notifications"),
        ("reconciliation_act_data", "reconciliation_act_data"),
        ("reconciliation_act", "reconciliation_act"),
        ("send_account_number", "send_account_number"),
        ("getcontractsbalance_byphonenumber", "GetContractsBalance_ByPhoneNumber"),
        ("getcontractsbalance_bycontractid", "GetContractsBalance_ByContractID"),
        ("getcontractsinfo_by_phone", "GetContractsInfo_By_Phone"),
        ("getmdswarninginfo_bycontractid", "GetMDsWarningInfo_ByContractID"),
        ("getmdswarninginfo_bytel", "GetMDsWarningInfo_ByTel"),
        ("getmdinfo_by_contractidandnomenclaturecode", "GetMDInfo_By_ContractIDAndNomenclatureCode"),
        ("inputofreadingswithdot", "InputOfReadingsWithDot"),
        ("findallbycontractnumber", "FindAllByContractNumber"),
        ("/info", "info"),
    ]
    keys = []
    for pattern, key in pairs:
        if pattern in v:
            keys.append(key)
    keys = sorted(set(keys))
    # убираем «родительские» ключи, если распознан более точный вариант по фамилии
    if "contracts_by_number_lastname" in keys and "contracts_by_number" in keys:
        keys.remove("contracts_by_number")
    if "contracts_by_address_lastname" in keys and "contracts_by_address" in keys:
        keys.remove("contracts_by_address")
    return "; ".join(keys) if keys else "—"


def derive_jsonpath(agent, slots_str, method):
    slots = slot_list(slots_str)
    if not slots:
        return "—"
    mapping = AGENT_SLOT_PATHS.get(agent) or {}
    if not mapping:
        return "—"
    parts = []
    for s in slots:
        if s in mapping:
            parts.append("%s = %s" % (s, mapping[s]))
    return "; ".join(parts) if parts else "—"


def derive_status(method, description):
    desc = (description or "").lower()
    m = (method or "").lower()
    if "агент уже существует (дополнить)" in desc:
        return "Реализовано (дополнить)"
    if "агент уже существует" in desc:
        return "Реализовано"
    if m == "не требуется":
        return "Не требуется"
    if m.startswith("вне api"):
        return "Не требуется (вне API)"
    if m.startswith("метод уточняется") or m.startswith("уточнить"):
        return "Уточнить у заказчика"
    return "Готово к разработке"


def enrich_row(sheet, row):
    section, question, id_article, request, method, fields, slots, agent, description = row
    method_key = derive_method_key(method)
    jsonpath = derive_jsonpath(agent, slots, method)
    status = derive_status(method, description)
    return [section, question, id_article, request, method, method_key, fields,
            jsonpath, slots, agent, status, description]


def collect_enriched():
    items = []
    for row in IDENT_ROWS:
        items.append(("Идентификация клиента",) + tuple(enrich_row("Идентификация клиента", row)))
    for title, rows in SCRIPT_ROWS.items():
        for row in rows:
            items.append((title,) + tuple(enrich_row(title, row)))
    return items


def style_header(ws, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = WRAP_C
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)), ws.max_row)


def method_fill(value):
    v = (value or "").lower()
    if "уточняется" in v or v.startswith("уточнить"):
        return FILL_ASK
    if "soap" in v:
        return FILL_SOAP
    if v == "не требуется" or v.startswith("вне api"):
        return FILL_NONE
    return FILL_REST


def write_rows(ws, sheet, rows):
    for i, row in enumerate(rows, start=1):
        values = [i] + enrich_row(sheet, row) + [
            FALLBACK_OPERATOR if row[7] and row[7] != "—" else "—"]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=i + 1, column=c, value=v)
            cell.border = BORDER
            cell.alignment = WRAP_C if c == 1 else WRAP
            if c == 6:
                cell.fill = method_fill(v)


def add_script_sheet(wb, title, rows):
    ws = wb.create_sheet(title)
    style_header(ws, HEADER_SCRIPT, WIDTHS_SCRIPT)
    write_rows(ws, title, rows)
    return ws


METHODS_META = {}


def init_methods_meta():
    for row in METHODS_ROWS:
        if row[0] not in METHODS_META:
            METHODS_META[row[0]] = row


METHOD_USAGE_EXTRA = {
    "disconnection_report_info": "Идентификация клиента (получение house_type)",
    "notifications": "Скрипт ОТКЛЮЧЕНИЕ ТСН — Жалоба на отсутствие оповещения (disabling_tsn + notifications)",
    "info": "Идентификация клиента (карточка ЛС); Скрипт ДИСТАНЦИОННЫЕ СЕРВИСЫ (регистрация в ЛК)",
}


def compute_usage(items):
    usage = {}
    for it in items:
        mkey = it[6]
        for k in [x.strip() for x in (mkey or "").split(";") if x.strip() and x != "—"]:
            usage.setdefault(k, {})
            usage[k][it[0]] = usage[k].get(it[0], 0) + 1
    return usage


def format_usage(name, usage):
    parts = []
    for sheet, cnt in sorted(usage.get(name, {}).items(), key=lambda x: -x[1]):
        parts.append("%s (%d)" % (sheet, cnt))
    extra = METHOD_USAGE_EXTRA.get(name)
    if extra:
        parts.append(extra)
    return "; ".join(parts) if parts else "—"


def add_methods_sheet(wb, rows, usage):
    ws = wb.create_sheet("Методы API")
    style_header(ws, HEADER_METHODS, WIDTHS_METHODS)
    for i, row in enumerate(rows, start=1):
        values = [i] + list(row) + [format_usage(row[0], usage)]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=i + 1, column=c, value=v)
            cell.border = BORDER
            cell.alignment = WRAP_C if c in (1, 3) else WRAP
            if c == 3:
                cell.fill = method_fill(v)
    return ws


def build_slots_sheet(wb, items):
    ws = wb.create_sheet("Слоты")
    style_header(ws, HEADER_SLOTS, WIDTHS_SLOTS)
    fills = {slot: [] for slot in SLOTS_INFO}
    uses = {slot: [] for slot in SLOTS_INFO}
    for it in items:
        sheet, section, question, _, _, _, _, _, _, slots, agent, _, _ = it
        label = scenario_label(sheet, section, question)
        for s in slot_list(slots):
            if s in fills:
                fills[s].append(label)
        for s in AGENT_INPUT_SLOTS.get(agent, []):
            if s in uses:
                uses[s].append(label)
    r = 1
    for slot, (typ, desc) in SLOTS_INFO.items():
        r += 1
        vals = [slot, typ, desc,
                "; ".join(fills[slot]) if fills[slot] else "—",
                "; ".join(uses[slot]) if uses[slot] else "—"]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = WRAP_C if c == 1 else WRAP
    return ws


def agent_status(statuses):
    order = ["Реализовано (дополнить)", "Реализовано", "Уточнить у заказчика",
             "Готово к разработке", "Не требуется (вне API)", "Не требуется"]
    for s in order:
        if s in statuses:
            return s
    return "; ".join(sorted(statuses)) if statuses else "—"


def build_union(agent, info):
    keys = info["keys"]
    rows = info["rows"]
    if not keys:
        return "—"
    if len(keys) == 1:
        k = list(keys)[0]
        if len(rows) > 1:
            meta = METHODS_META.get(k)
            if meta:
                url, fields = meta[2], meta[5]
                return ("Один вызов %s (%s) возвращает: %s. Покрывает все %d вопроса(ов) из "
                        "колонки «Вопросы» — реализовать ОДИН агент." % (url, k, fields, len(rows)))
        return "—"
    return "Вопросы требуют разных вызовов: %s" % "; ".join(sorted(keys))


def build_agents_sheet(wb, items):
    ws = wb.create_sheet("Агенты")
    style_header(ws, HEADER_AGENTS, WIDTHS_AGENTS)
    agents = {}
    for it in items:
        sheet, section, question, _, _, _, mkey, _, _, _, agent, status, _ = it
        if not agent or agent == "—":
            continue
        info = agents.setdefault(agent, {"rows": [], "statuses": set(), "keys": set()})
        info["rows"].append((sheet, section, question))
        info["statuses"].add(status)
        for k in [x.strip() for x in (mkey or "").split(";") if x.strip() and x != "—"]:
            info["keys"].add(k)
    r = 1
    for agent in sorted(agents):
        info = agents[agent]
        keys = "; ".join(sorted(info["keys"])) if info["keys"] else "—"
        questions = "\n".join(scenario_label(s, sec, q) for s, sec, q in info["rows"])
        r += 1
        vals = [agent, agent_status(info["statuses"]), keys, questions,
                build_union(agent, info), FALLBACK_OPERATOR]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = WRAP_C if c == 1 else WRAP
    return ws


# Колонки строки: лист, вопрос, id статьи, что запросить, метод, поля ответа, слоты, id агента, описание агента
IDENT_ROWS = [
    ("—", "Клиент идентифицирован по номеру звонящего: найти ЛС по номеру телефона (может быть несколько ЛС — уточнять адрес)",
     "", "Поиск лицевых счетов по номеру входящего телефона",
     "REST GET /contracts/by_phone/{phone}",
     "UserId; WebProperties.ContractNo; WebProperties.Address; WebProperties.IsRegistered",
     "user_id; contract_no; address",
     "identification",
     "Агент уже существует. Вызвать contracts_by_phone по номеру звонящего; при 1 ЛС — подтвердить адрес, при нескольких — уточнить адрес; заполнить слоты user_id, contract_no, address; запросить house_type через disconnection_report_info и положить в слот house_type; при отсутствии совпадений — перевод на оператора (final_answer='error')."),
    ("—", "Проверка ФИО владельца ЛС при идентификации по адресу/телефону",
     "", "Проверка «на кого оформлен лицевой счет» по карточке ЛС",
     "REST GET /info/{user_id}",
     "account.number; account.owner; house.address",
     "contract_no; address; owner_fio",
     "get_client_info",
     "Выполнить GET /info/{user_id}; положить в слоты contract_no (account.number), address (house.address), owner_fio (account.owner), status (account.status); в final_answer вернуть признак успешной идентификации владельца."),
    ("—", "Поиск ЛС по номеру (клиент назвал номер ЛС)",
     "", "Найти ЛС по номеру и проверить ФИО/адрес",
     "REST GET /contracts/by_number/{contractNo}",
     "UserId; WebProperties.ContractNo; WebProperties.Address",
     "user_id; contract_no; address",
     "contracts_by_number",
     "Выполнить GET /contracts/by_number/{contractNo}; при совпадении заполнить слоты user_id, contract_no, address; при отсутствии — final_answer='not_found' или перевод на оператора."),
    ("—", "Поиск ЛС по адресу и ФИО (клиент назвал адрес и ФИО)",
     "", "Найти ЛС по адресу и фамилии владельца",
     "REST GET /contracts/by_address_lastname",
     "UserId; WebProperties.ContractNo",
     "user_id; contract_no",
     "contracts_by_address_lastname",
     "Выполнить GET /contracts/by_address_lastname (lastName, cityName, streetName, houseName, flatName); при совпадении заполнить слоты user_id, contract_no; иначе — перевод на оператора."),
    ("—", "Получение всех ЛС по адресу (все квартиры дома)",
     "", "Список ЛС по адресу для выбора нужного",
     "REST GET /contracts/by_address?search_all_flats=true",
     "UserId[]; WebProperties.ContractNo[]; WebProperties.Address[]",
     "user_id; contract_no",
     "contracts_by_address",
     "Выполнить GET /contracts/by_address?search_all_flats=true; получить список ЛС по адресу; заполнить слоты user_id, contract_no; при нескольких ЛС — вернуть список для выбора (final_answer с количеством)."),
    ("—", "Проверка ЛС по номеру и фамилии",
     "", "Найти ЛС по номеру и фамилии владельца",
     "REST GET /contracts/by_number_lastname/{contractNo}/{lastName}",
     "UserId; WebProperties.ContractNo",
     "user_id; contract_no",
     "contracts_by_number_lastname",
     "Выполнить GET /contracts/by_number_lastname/{contractNo}/{lastName}; при совпадении заполнить слоты user_id, contract_no; иначе — перевод на оператора."),
    ("—", "Открытие карточки ЛС: статус, баланс, услуги на расчетах, приборы учета, дом",
     "", "Полная карточка лицевого счета",
     "REST GET /info/{user_id}",
     "account.number; account.status; account.balance; account.owner; account.is_owner_registered; account.devices[]; account.accounts[]; house.address",
     "contract_no; address; balance; is_owner_registered; service; device_status",
     "get_client_info",
     "Выполнить GET /info/{user_id}; заполнить слоты: contract_no, address, balance (account.balance), is_owner_registered, service (account.accounts[]), device_status (account.devices[]); в final_answer вернуть статус ЛС."),
    ("—", "(резервный канал) Поиск ЛС по номеру телефона через SOAP-шину IVR",
     "", "Поиск ЛС по телефону (дублирующий способ)",
     "SOAP IVR.asmx GetContractsInfo_By_Phone",
     "ContractInfo{ID; No; Adress; AbonentName; Status; DateUpdate}",
     "user_id; contract_no; address",
     "soap_contracts_by_phone",
     "SOAP-запрос GetContractsInfo_By_Phone (PhoneNumber); из ContractInfo заполнить слоты user_id (ID), contract_no (No), address (Adress), owner_fio (AbonentName); резервный канал идентификации."),
]


SCRIPT_ROWS = {
    "Скрипт ОТКЛЮЧЕНИЕ ЭЭ": [
        ("Уточнение причины отключения ЭЭ",
         "Клиент не идентифицирован / хочет узнать инфо по другому адресу: как проверять отключения по адресу? (запрос disconnections работает по user_id)",
         "", "Поиск ЛС по адресу, чтобы получить user_id для проверки отключений",
         "REST GET /contracts/by_address",
         "UserId[]; WebProperties.ContractNo[]; WebProperties.Address[]",
         "user_id; contract_no; address",
         "contracts_by_address",
         "Выполнить GET /contracts/by_address; получить список ЛС по адресу; заполнить слоты user_id, contract_no; при нескольких ЛС — предложить выбрать адрес (final_answer с вариантами)."),
        ("Уточнение причины отключения ЭЭ",
         "Запрос на получение информации об отключении на сайте",
         "", "Проверить наряд/отключение и задолженность по ЛС",
         "REST GET /{user_id}/disconnections_electro",
         "disconnection{text; type; work_type; date_start; date_end}; debt",
         "disconnection; debt",
         "disabling_electro",
         "Агент уже существует. Вызвать disabling_electro по user_id (GET /disconnections_electro); положить в слоты disconnection, debt, house_type; заполнить final_answer: 1 — есть отключение (текст отключения в слот disconnection), 2 — есть долг, 3 — нет ни того, ни другого."),
        ("Уточнение причины отключения ЭЭ",
         "Запрос на наличие оповещения на портале КЦ (поиск по произведенной рассылке/обзвону)",
         "", "Оповещение в РМО (виджет Телемаркетинг) — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "notification_kc_check",
         "Метод уточняется у заказчика. Проверить наличие оповещения в портале КЦ (рассылка/обзвон, виджет Телемаркетинг); заполнить final_answer: 1 — оповещение было, 0 — не было."),
        ("Уточнение причины отключения ЭЭ",
         "Запрос на получение информации о отправке Push-уведомлений в ЛК на сайте",
         "", "Push-уведомления в ЛК — метод уточняется у заказчика (есть ли интеграция с сайтом)",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "notification_push_check",
         "Метод уточняется у заказчика (интеграция с сайтом/ЛК). Проверить отправку Push-уведомления в ЛК ФЛ; заполнить слоты и final_answer при наличии данных."),
        ("Проверка наличия наряда",
         "Клиент не идентифицирован — метод поиска по адресу не реализован",
         "", "Поиск ЛС по адресу и ФИО владельца",
         "REST GET /contracts/by_address_lastname",
         "UserId; WebProperties.ContractNo",
         "user_id; contract_no",
         "contracts_by_address_lastname",
         "Выполнить GET /contracts/by_address_lastname (адрес + ФИО); при совпадении заполнить слоты user_id, contract_no; при отсутствии совпадений — перевод на оператора."),
        ("Проверка наличия наряда",
         "Проверка наличия наряда на ограничение электроэнергии",
         "", "Получить наряд/ограничение и задолженность по ЛС",
         "REST GET /{user_id}/disconnections_electro",
         "disconnection{text; work_type; date_start; date_end}; debt",
         "disconnection; debt",
         "disabling_electro",
         "Агент уже существует. Вызвать disabling_electro по user_id; проверить наличие наряда/отключения и долга; заполнить слоты disconnection, debt; final_answer: 1 — есть наряд/отключение, 2 — долг, 3 — нет наряда и долга."),
        ("Проверка наличия наряда",
         "Запрос изменения данных в лицевом счете",
         "", "Изменение данных/контактного телефона в ЛС — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "client_data_change",
         "Метод уточняется у заказчика. Реализовать изменение данных/контактного телефона в ЛС; в final_answer вернуть результат изменения."),
        ("Проверка наличия наряда",
         "Запрос получения ограничения по лицевому счету",
         "", "Получить текущее ограничение и задолженность",
         "REST GET /{user_id}/disconnections_electro",
         "disconnection; debt",
         "disconnection; debt",
         "disabling_electro",
         "Агент уже существует. Вызвать disabling_electro по user_id; получить текущее ограничение и задолженность; заполнить слоты disconnection, debt; final_answer по условиям."),
        ("Вопросы по наряду",
         "Запрос получения задолженности по лицевому счету",
         "", "Получить текущую задолженность по ЛС",
         "REST GET /{user_id}/info",
         "account.balance",
         "balance",
         "get_balance",
         "Выполнить GET /info/{user_id}; положить в слот balance (account.balance); в final_answer озвучить размер задолженности."),
        ("Вопросы по наряду",
         "Запрос получения последней поступившей оплаты и способа поступления денежных средств",
         "", "Получить последние платежи и способ поступления",
         "REST GET /{user_id}/transactions",
         "transactions[].transactions[]{payment; date}",
         "last_payment",
         "get_transactions",
         "Выполнить GET /transactions/{user_id}; из последней оплаты заполнить слот last_payment (сумма, дата, способ); в final_answer вернуть текст о последней оплате."),
        ("нет плана_аварии_наряда ЧД",
         "Запрос оформления заявки в КСРТ (авария, частный дом)",
         "", "Оформить аварийную заявку в КСРТ",
         "REST POST /{user_id}/disconnection_report_info",
         "id; status",
         "—",
         "ksrt_report_create",
         "Выполнить POST /disconnection_report_info/{user_id} с данными заявки (адрес/ЛС, причина, контакты); в final_answer вернуть статус оформления; id заявки положить в слот."),
        ("Садоводство",
         "Запрос на получение плановых работ",
         "", "Получить плановые отключения по ЛС",
         "REST GET /{user_id}/disconnections_electro",
         "disconnection{work_type; date_start; date_end}",
         "disconnection",
         "disabling_electro",
         "Агент уже существует. Вызвать disabling_electro по user_id; из disconnection выбрать плановые работы (work_type); заполнить слот disconnection; в final_answer озвучить плановые отключения."),
        ("Садоводство",
         "Запрос на получение текущих отключений",
         "", "Получить текущие отключения по ЛС",
         "REST GET /{user_id}/disconnections_electro",
         "disconnection",
         "disconnection",
         "disabling_electro",
         "Агент уже существует. Вызвать disabling_electro по user_id; получить текущие отключения; заполнить слот disconnection; final_answer по условиям."),
        ("Опора",
         "Запрос на получение сетевой организации по адресу",
         "", "Получить сетевую организацию по ЛС",
         "REST GET /{user_id}/network_organization_data",
         "Name; Address; Contacts",
         "network_org",
         "network_org",
         "Выполнить GET /network_organization_data/{user_id}; заполнить слот network_org (Name, Address, Contacts); в final_answer озвучить наименование и контакты сетевой организации."),
        ("Качество ЭЭ",
         "Проверка наличия жалобы на низкое напряжение в CRM и даты её оформления",
         "", "Поиск жалобы на качество напряжения — метод уточняется у заказчика (такого запроса нет)",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "voltage_complaint_check",
         "Метод уточняется у заказчика. Проверить наличие жалобы на низкое/высокое напряжение в CRM и дату её оформления; заполнить слоты и final_answer."),
        ("не работает пульт",
         "Оформление заявки в ИЭСК (не работает пульт / отсутствует фаза / мигает свет)",
         "", "Оформить аварийную заявку в ИЭСК",
         "REST POST /{user_id}/disconnection_report_info",
         "id; status",
         "—",
         "ksrt_report_create",
         "Выполнить POST /disconnection_report_info/{user_id}; оформить заявку в ИЭСК (причина: не работает пульт / отсутствует фаза); в final_answer вернуть статус; id заявки в слот."),
        ("Проверка наличия наряда",
         "«Я оплатил долг по электроэнергии, пожалуйста, отмените наряд на ограничение»",
         "", "Отменить заявку КСРТ",
         "REST POST /{user_id}/cancel_ksrt",
         "id; status",
         "—",
         "ksrt_cancel",
         "Выполнить POST /cancel_ksrt/{user_id}; отменить заявку КСРТ после оплаты долга; в final_answer вернуть статус отмены."),
    ],

    "Скрипт ДОГОВОРНАЯ РАБОТА": [
        ("Идентификация клиента",
         "Идентификация по номеру звонящего / номеру ЛС / адресу и ФИО",
         "", "Общая идентификация (см. лист «Идентификация клиента»)",
         "REST GET /contracts/by_phone (или /by_number, /by_address_lastname)",
         "UserId; WebProperties.ContractNo; WebProperties.Address",
         "user_id; contract_no; address",
         "identification",
         "Агент уже существует. Общая идентификация по телефону/ЛС/адресу (см. лист «Идентификация клиента»); заполнить слоты user_id, contract_no, address."),
        ("Открытие ЛС, новый собственник",
         "Запрос возвращающий дату открытия ЛС",
         "", "Получить дату открытия ЛС — актуальность поля в /info уточняется у заказчика",
         "метод уточняется у заказчика (проверить /info)",
         "account.number; дата открытия ЛС (уточняется)",
         "contract_no",
         "get_contract_open_date",
         "Метод уточняется у заказчика. Получить дату открытия ЛС (проверить наличие поля в /info или отдельный метод); заполнить слот contract_open_date; в final_answer озвучить дату."),
        ("Открытие ЛС, новый собственник",
         "«Лицевой счет на (ФИО) открыт с (дата)» — проверка",
         "", "Найти ЛС по адресу и ФИО нового собственника",
         "REST GET /contracts/by_address_lastname",
         "UserId; WebProperties.ContractNo",
         "user_id; contract_no",
         "contracts_by_address_lastname",
         "Найти ЛС по адресу и ФИО нового собственника; заполнить слоты user_id, contract_no; подтвердить владельца (owner_fio)."),
        ("Открытие ЛС, новый собственник",
         "Проверка наличия ЛС по адресу и фамилии (новый собственник)",
         "", "Найти ЛС по адресу и фамилии",
         "REST GET /contracts/by_address_lastname",
         "UserId; WebProperties.ContractNo",
         "user_id; contract_no",
         "contracts_by_address_lastname",
         "Выполнить GET /contracts/by_address_lastname; при совпадении заполнить слоты user_id, contract_no; при отсутствии — перевод на оператора."),
        ("Расторжение договора",
         "Идентификация клиента по номеру телефона (номеру ЛС) → GetUserInfo",
         "", "Получить карточку ЛС (GetUserInfo)",
         "REST GET /{user_id}/info",
         "account.status; account.owner; account.number",
         "user_id; contract_no",
         "get_client_info",
         "Выполнить GET /info/{user_id} (GetUserInfo); заполнить слоты user_id, contract_no, status, owner_fio; проверить возможность расторжения по status."),
        ("Расторжение договора",
         "Проверка наличия ЛС по адресу и фамилии (закрытие ЛС)",
         "", "Найти ЛС по адресу и фамилии",
         "REST GET /contracts/by_address_lastname",
         "UserId; WebProperties.ContractNo",
         "user_id; contract_no",
         "contracts_by_address_lastname",
         "Выполнить GET /contracts/by_address_lastname; заполнить слоты user_id, contract_no; при отсутствии совпадений — перевод на оператора."),
        ("ЛС выключен",
         "Запрос связки ЛС или получение всех ЛС по адресу",
         "", "Получить все ЛС по адресу (все квартиры дома)",
         "REST GET /contracts/by_address?search_all_flats=true",
         "UserId[]; WebProperties.ContractNo[]",
         "user_id; contract_no",
         "contracts_by_address",
         "Получить все ЛС по адресу (search_all_flats=true); заполнить слоты user_id, contract_no; вывести список для выбора."),
        ("Изменения",
         "Запрос внесение изменений в ЛС",
         "", "Внесение изменений в данные ЛС — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "client_data_change",
         "Метод уточняется у заказчика. Внесение изменений в ЛС (ФИО/адрес/контакты); в final_answer вернуть результат изменения."),
        ("Номер ЛС",
         "Проверка наличия ЛС по номеру",
         "", "Найти ЛС по номеру",
         "REST GET /contracts/by_number/{contractNo}",
         "UserId; WebProperties.Address",
         "user_id; address",
         "contracts_by_number",
         "Выполнить GET /contracts/by_number/{contractNo}; при совпадении заполнить слоты user_id, address; при отсутствии — перевод на оператора."),
        ("Переход МКД",
         "Запрос получение договора по ЛС",
         "", "Проверить наличие договора/ЛС по номеру",
         "REST GET /contracts/by_number/{contractNo}",
         "UserId; WebProperties.ContractNo",
         "user_id; contract_no",
         "contracts_by_number",
         "Проверить наличие договора/ЛС по номеру; заполнить слоты user_id, contract_no; в final_answer вернуть признак наличия договора."),
        ("Электронная квитанция",
         "Запрос: включение печати бумажной квитанции",
         "", "Переключение способа выставления квитанций (бумажная/электронная) — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "billing_receipt_mode",
         "Метод уточняется у заказчика. Переключить способ выставления квитанции (бумажная/электронная) по ЛС; в final_answer вернуть результат."),
        ("Электронная квитанция",
         "Запрос возвращающий признак «Выставлять квитанции» = Да",
         "", "Получить признак выставления квитанций — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "billing_receipt_mode",
         "Метод уточняется у заказчика. Получить признак «Выставлять квитанции»; положить в слот receipt_paper_flag (Да/Нет)."),
    ],

    "Скрипт ПРИБОРЫ УЧЁТА": [
        ("Идентификация клиента",
         "Идентификация по адресу и ФИО владельца",
         "", "Общая идентификация (см. лист «Идентификация клиента»)",
         "REST GET /contracts/by_address_lastname",
         "UserId; WebProperties.ContractNo",
         "user_id; contract_no",
         "identification",
         "Агент уже существует. Общая идентификация по адресу и ФИО (см. лист «Идентификация клиента»); заполнить слоты user_id, contract_no."),
        ("опломбир пу ээ",
         "Проверить в ЛС наличие прибора учета ЭЭ на расчетах",
         "", "Получить приборы учета ЛС и их статусы",
         "REST GET /{user_id}/devices",
         "devices[]{service_name; is_installed; status; accepts_readings; readings_accept_type}",
         "device_status",
         "devices_check",
         "Выполнить GET /devices/{user_id}; заполнить слоты device_status, last_reading, next_check по выбранной услуге; в final_answer вернуть наличие прибора на расчетах."),
        ("опломбир пу ээ",
         "Проверить наличие наряда (отключение по задолженности)",
         "", "Получить наряд/отключение и задолженность",
         "REST GET /{user_id}/disconnections_electro",
         "disconnection; debt",
         "disconnection; debt",
         "disabling_electro",
         "Агент уже существует. Проверить наряд на отключение по задолженности (GET /disconnections_electro); заполнить слоты disconnection, debt; final_answer 1/2/3."),
        ("опломбир пу ээ / если нет пу в АСРН / снятие пломбы МКД",
         "Оформление заявки на опломбировку / обследование / проверку прибора учета",
         "", "Создание заявки на вызов специалиста — метод уточняется у заказчика (GET specialist_requests только читает заявки)",
         "метод уточняется у заказчика (создание specialist_requests)",
         "id; status; назначенная дата",
         "—",
         "specialist_request_create",
         "Метод уточняется у заказчика (GET /specialist_requests только читает заявки). Создать заявку на опломбировку/обследование/проверку ПУ; в final_answer вернуть статус; id заявки в слот."),
        ("дата опломбир / дата обследов",
         "Проверка наличия заявки и назначенной даты",
         "", "Получить заявки на специалиста и их статусы",
         "REST GET /{user_id}/specialist_requests",
         "requests[]{id; request_date; title; status}",
         "request_status",
         "specialist_requests_check",
         "Выполнить GET /specialist_requests/{user_id}; заполнить слот request_status (id, request_date, title, status); в final_answer вернуть наличие заявки и назначенную дату."),
        ("Сверка пок",
         "Запрос последних показаний по выбранному прибору учета",
         "", "Получить последние показания по прибору",
         "REST GET /{user_id}/devices",
         "devices[].last_reading",
         "last_reading",
         "devices_check",
         "Выполнить GET /devices/{user_id}; получить последние показания по выбранному прибору (last_reading); положить в слот last_reading; в final_answer озвучить показания и дату."),
        ("Магеллан / передача пок по кон.центр",
         "Проверить наличие показаний в лицевом счете (последние показания и дата)",
         "", "Получить историю показаний по прибору",
         "REST GET /{user_id}/devices/history",
         "history[]{value; date}; last_reading",
         "readings_history",
         "device_history_get",
         "Выполнить GET /devices/history/{user_id}; заполнить слот readings_history (value, date), last_reading; в final_answer озвучить последние показания и дату внесения."),
        ("Обновление ЛК / Передача показаний",
         "Запрос на получении истории передачи показаний в ЛС",
         "", "Получить историю передачи показаний и статус",
         "REST GET /{user_id}/devices/history",
         "history[]{value; date; status_transfer}; last_reading",
         "readings_history",
         "device_history_get",
         "Выполнить GET /devices/history/{user_id}; заполнить слот readings_history (value, date, status_transfer); в final_answer вернуть приняты ли показания и когда."),
        ("поверка",
         "Срок поверки (МПИ) прибора учета",
         "", "Получить дату следующей поверки ПУ",
         "REST GET /{user_id}/devices",
         "devices[].next_check",
         "next_check",
         "device_verification_check",
         "Выполнить GET /devices/{user_id} (next_check) или SOAP GetMDsWarningInfo_ByContractID; заполнить слот next_check; в final_answer озвучить дату поверки/предупреждение об истечении МПИ."),
        ("Истечение МПИ ПУ ЭЭ / опломбир ипу отопления",
         "Идентификация. Проверка наличия услуги на расчетах и типа строения",
         "", "Получить услуги ЛС на расчетах и тип дома",
         "REST GET /{user_id}/info",
         "account.accounts[]; house_type (из /disconnection_report_info)",
         "service; house_type",
         "get_client_info",
         "Выполнить GET /info/{user_id}; заполнить слот service (account.accounts[]); получить house_type из disconnection_report_info и положить в слот house_type."),
        ("интелл ПУ ЭЭ нет показаний / ПУ РиМ",
         "Интеллектуальный прибор учета: проверка автоматической передачи показаний",
         "", "Проверить тип ПУ и его статус",
         "REST GET /{user_id}/devices",
         "devices[]{is_smart; is_installed; status}",
         "is_smart; device_status",
         "devices_check",
         "Выполнить GET /devices/{user_id}; проверить is_smart и статус прибора; заполнить слоты is_smart, device_status; в final_answer объяснить про автоматическую передачу показаний."),
        ("не прин через сайт без рег / лицевой счет выключен / если нет пу в АСРН",
         "Проверка прибора и статуса снятия с расчетов",
         "", "Проверить статус ПУ и наличие услуги на расчетах",
         "REST GET /{user_id}/devices; GET /{user_id}/info",
         "devices[].status; devices[].accepts_readings; account.accounts[]",
         "device_status; service",
         "devices_check",
         "Выполнить GET /devices/{user_id} и GET /info/{user_id}; проверить статус ПУ (accepts_readings, status) и услуги на расчетах; заполнить слоты device_status, service; final_answer — причина отказа приема показаний."),
        ("снятие пломбы Част ТП",
         "Смотрим сетевую организацию",
         "", "Получить сетевую организацию по ЛС",
         "REST GET /{user_id}/network_organization_data",
         "Name; Contacts",
         "network_org",
         "network_org",
         "Выполнить GET /network_organization_data/{user_id}; заполнить слот network_org (Name, Contacts); в final_answer озвучить наименование сетевой организации."),
        ("Установка пу ЭЭ домовладение / Установка-замена ээ МКД",
         "Смотрим есть ли прибор учета в ЛС и его статус в /info.account.devices.readings_accept_type",
         "", "Получить статус прибора учета из карточки ЛС",
         "REST GET /{user_id}/info",
         "account.devices[]{service_name; status; readings_accept_type}",
         "device_status",
         "get_client_info",
         "Выполнить GET /info/{user_id}; из account.devices.readings_accept_type определить статус прибора; заполнить слот device_status; final_answer по условиям (на расчетах / снят)."),
        ("поверка (ГВС, ХВС, Отопления)",
         "Посмотреть дату поверки (сообщить клиенту)",
         "", "Получить дату следующей поверки ПУ",
         "REST GET /{user_id}/devices (next_check) или SOAP GetMDsWarningInfo_ByContractID",
         "devices[].next_check; MDWarningInfo.NextVerificationDeadline",
         "next_check",
         "device_verification_check",
         "Выполнить GET /devices/{user_id} (next_check) или SOAP GetMDsWarningInfo_ByContractID; заполнить слот next_check; в final_answer озвучить дату поверки."),
    ],

    "Скрипт СЕРВИСНЫЙ ЦЕНТР": [
        ("Нет технологическог подключения",
         "Нужна интеграция для проверки зон обслуживания — с какой системой?",
         "", "Определение зоны обслуживания сетевой организации по адресу — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "service_zone_check",
         "Метод уточняется у заказчика. Определить зону обслуживания сетевой организации по адресу; заполнить слот network_org/service_zone; в final_answer вернуть наименование зоны."),
        ("Готовность договора",
         "Нужна интеграция — проверка готовности договора технологического присоединения",
         "", "Проверка готовности/статуса договора ТП — метод уточняется у заказчика (источник: ЛК ИЭСК)",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "tp_contract_ready_check",
         "Метод уточняется у заказчика (источник: ЛК ИЭСК). Проверить готовность/статус договора техприсоединения; заполнить слот tp_status; в final_answer озвучить статус."),
        ("Увеличение мощности (эле подкл)",
         "Восстановление документов о технологическом присоединении",
         "", "Заявка через ЛК ИЭСК; для других сетевых организаций — метод уточняется у заказчика",
         "вне API (ЛК ИЭСК)",
         "—",
         "—",
         "tp_docs_restore",
         "Метод уточняется у заказчика (для сетевых организаций вне ИЭСК). Заявка на восстановление документов ТП через ЛК ИЭСК; в final_answer — инструкция и способ подачи заявки."),
        ("Подключение объекта микрогенерации",
         "Отслеживание статуса заявки на техприсоединение (микрогенерация)",
         "", "Статус заявки в ЛК ИЭСК — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "tp_request_status",
         "Метод уточняется у заказчика. Получить статус заявки на техприсоединение (микрогенерация) в ЛК ИЭСК; заполнить слот tp_status; final_answer — статус заявки."),
        ("Документы",
         "Отправка ссылки на сайт клиенту на телефон (СМС/мессенджер)",
         "", "Отправка SMS/сообщения со ссылкой — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "send_link_message",
         "Метод уточняется у заказчика. Отправка SMS/сообщения со ссылкой на сайт клиенту; в final_answer вернуть результат отправки."),
    ],

    "Скрипт СПРАВОЧНАЯ ИНФОРМАЦИЯ": [
        ("Контакты (почта, реквизиты)",
         "Назовите реквизиты компании: ИНН, ОГРН, платежные реквизиты и т.д.",
         "", "Статическая справочная информация (данные из статей/настроек агента) — интеграция не требуется",
         "не требуется",
         "—",
         "—",
         "—",
         "Агент не требуется: данные статические (ИНН, ОГРН, платежные реквизиты), хранятся в статьях или agentSettings."),
        ("Уличное освещение / служба Сигнал / Сотрудничество",
         "Консультации по справочной информации",
         "", "Справочные данные без обращения к внешним системам",
         "не требуется",
         "—",
         "—",
         "—",
         "Агент не требуется: ответы формируются из статей/настроек без обращений к API."),
    ],

    "Скрипт ОТКЛЮЧЕНИЕ ТСН": [
        ("Отключение гор.водосн-я,теплосн",
         "Идентификация по номеру звонящего / номеру ЛС / адресу и ФИО",
         "", "Общая идентификация (см. лист «Идентификация клиента»); метод проверки по адресу не реализован — перевод к оператору",
         "REST GET /contracts/by_phone (или /by_number, /by_address_lastname)",
         "UserId; WebProperties.ContractNo; WebProperties.Address",
         "user_id; contract_no; address",
         "identification",
         "Агент уже существует. Общая идентификация (см. лист «Идентификация клиента»); при отсутствии совпадений и отсутствии метода проверки по адресу — перевод к оператору."),
        ("Отключение гор.водосн-я,теплосн",
         "Проверка отключения ГВС/теплоснабжения и задолженности (агент disabling_tsn)",
         "", "Вызов JS-агента disabling_tsn: передать идентификатор ЛС в поле user_id; из ответа взять disconnection и debt",
         "REST GET /{user_id}/disconnections_heat",
         "disconnection{text; work_type; date_start; date_end}; debt",
         "disconnection; debt",
         "disabling_tsn",
         "Агент уже существует. Вызвать disabling_tsn, передав ЛС в user_id (GET /disconnections_heat); проверить наличие отключения и наряда; положить доступные значения в слоты (house_type и др.); заполнить final_answer: 1 — есть отключение (текст отключения в слот disconnection, при необходимости дату окончания работ), 2 — есть долг, 3 — нет отключения и долга."),
        ("Отключение гор.водосн-я,теплосн",
         "Проверка плановых работ / наряда: сверить дату окончания работ (агент disabling_tsn_plan)",
         "", "Вызов JS-агента disabling_tsn_plan: получить disconnection.date_end и сравнить с текущей датой",
         "REST GET /{user_id}/disconnections_heat",
         "disconnection{text; date_start; date_end; work_type}",
         "disconnection",
         "disabling_tsn_plan",
         "Агент уже существует. Вызвать disabling_tsn_plan по user_id; получить disconnection.date_end; если дата окончания работ в будущем — озвучить её (final_answer=1), если уже прошла — перевод на оператора; текст отключения в слот disconnection."),
        ("Жалоба на отсутствие оповещения",
         "Проверка наличия уведомления в ответе (для письма на network@es.irkutskenergo.ru)",
         "", "Получить уведомление/отключение по ЛС (тип, дата начала и окончания)",
         "REST GET /{user_id}/disconnections_heat; GET /{user_id}/notifications",
         "disconnection{type; date_start; date_end}; notifications",
         "disconnection; notifications",
         "disabling_tsn",
         "Агент уже существует (дополнить). Вызвать disabling_tsn и notifications; проверить наличие уведомления; собрать данные для письма на network@es.irkutskenergo.ru (тема «Нет оповещения»): адрес, ЛС, ФИО, телефон, услугу, период отключения (type, date_start, date_end); заполнить слоты disconnection, notifications."),
        ("Жалоба на отсутствие оповещения",
         "Получение ЛС и номера телефона при звонке — в каких слотах?",
         "", "Идентификация по номеру звонящего: слоты из входящего сообщения",
         "REST GET /contracts/by_phone",
         "UserId; WebProperties.ContractNo",
         "user_id; phone; contract_no",
         "identification",
         "Агент уже существует. Идентификация по номеру звонящего; из входящего сообщения взять номер телефона и положить в слот phone; заполнить user_id, contract_no."),
        ("Планы управляющей компании",
         "Проверка планов/аварии по адресу клиента → наличие наряда",
         "", "Получить отключения по ЛС (план/авария, наряд)",
         "REST GET /{user_id}/disconnections_heat",
         "disconnection; debt",
         "disconnection",
         "disabling_tsn",
         "Агент уже существует. Вызвать disabling_tsn по user_id; проверить планы/аварии (disconnections_heat) и наличие наряда; заполнить слоты disconnection, debt; final_answer по условиям."),
        ("Планы управляющей компании",
         "Жалоба в АСУСЭиРП",
         "", "Внешняя система АСУСЭиРП — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "asus_esrp_complaint",
         "Метод уточняется у заказчика. Оформление жалобы в АСУСЭиРП; заполнить слоты и final_answer по статусу жалобы."),
    ],

    "Скрипт ДИСТАНЦИОННЫЕ СЕРВИСЫ": [
        ("Идентификация клиента",
         "Идентификация по адресу и ФИО владельца",
         "", "Общая идентификация (см. лист «Идентификация клиента»)",
         "REST GET /contracts/by_address_lastname",
         "UserId; WebProperties.ContractNo",
         "user_id; contract_no",
         "identification",
         "Агент уже существует. Общая идентификация по адресу и ФИО (см. лист «Идентификация клиента»); заполнить слоты user_id, contract_no."),
        ("Номер ЛС",
         "Идентифицируем по номеру звонящего: найти ЛС и озвучить номер ЛС",
         "", "Поиск ЛС по номеру телефона",
         "REST GET /contracts/by_phone/{phone}",
         "UserId; WebProperties.ContractNo; WebProperties.Address",
         "user_id; contract_no",
         "identification",
         "Агент уже существует. Поиск ЛС по номеру звонящего (contracts_by_phone); при 1 ЛС — озвучить номер ЛС, при нескольких — уточнить адрес; заполнить слоты user_id, contract_no."),
        ("Номер ЛС",
         "Как получить список жильцов по адресу, какое API?",
         "", "Список жильцов/зарегистрированных по адресу — метод уточняется у заказчика (нет в REST API)",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "residents_list",
         "Метод уточняется у заказчика. Получить список жильцов/зарегистрированных по адресу; заполнить слот residents; в final_answer перечислить жильцов."),
        ("Личный кабинет ФЛ",
         "Запросить номер ЛС и проверить статус регистрации в ЛК по ЛС (поле account.is_owner_registered)",
         "", "Получить признак регистрации в Личном кабинете",
         "REST GET /{user_id}/info",
         "account.is_owner_registered; account.number",
         "is_owner_registered",
         "get_client_info",
         "Выполнить GET /info/{user_id}; заполнить слот is_owner_registered (account.is_owner_registered), contract_no; в final_answer вернуть признак регистрации в ЛК."),
        ("Мобильное приложение ФЛ",
         "Запрос на сброс регистрации",
         "", "Сброс регистрации в ЛК/МП — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "registration_reset",
         "Метод уточняется у заказчика. Сброс регистрации в ЛК/МП; в final_answer вернуть результат сброса."),
        ("Передача показаний",
         "Запрос на получении истории передачи показаний в ЛС",
         "", "Получить историю показаний и статус передачи",
         "REST GET /{user_id}/devices/history",
         "history[]{value; date; status_transfer}; last_reading",
         "readings_history",
         "device_history_get",
         "Выполнить GET /devices/history/{user_id}; заполнить слот readings_history (value, date, status_transfer); в final_answer вернуть приняты ли показания и когда."),
        ("Передача показаний",
         "Я передал показания через ЛК/МП, когда вы их увидите? (проверка в АСРН)",
         "", "Проверить последние переданные показания в ЛС",
         "REST GET /{user_id}/devices/history",
         "last_reading; history[0].value; history[0].date",
         "last_reading",
         "device_history_get",
         "Выполнить GET /devices/history/{user_id}; проверить последние переданные показания (last_reading, история); заполнить слоты; в final_answer — статус и сроки появления в ЛС."),
        ("Ошибки в сервисах",
         "Как узнать, что услуга не отключена, что ИПУ не привязан к услуге, что прибор не снят с расчетов",
         "", "Получить статус услуги и приборов учета",
         "REST GET /{user_id}/devices; GET /{user_id}/info",
         "devices[].status; devices[].accepts_readings; account.accounts[]",
         "device_status; service",
         "devices_check",
         "Выполнить GET /devices/{user_id} и GET /info/{user_id}; проверить, что услуга не отключена, ИПУ привязан, ПУ не снят с расчетов (devices.status, accepts_readings, account.accounts[]); заполнить слоты device_status, service; final_answer — причина ошибки."),
        ("Ошибки в сервисах",
         "Как оформить заявку в службу поддержки? (интеграция с системой техподдержки)",
         "", "API получения/оформления заявки в службу поддержки — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "support_request_create",
         "Метод уточняется у заказчика (интеграция с системой техподдержки). Оформить заявку в службу поддержки по ЛС; в final_answer вернуть статус; id заявки в слот."),
        ("Ошибки в сервисах",
         "Нужно API получения заявки в службу поддержки",
         "", "Заявки техподдержки — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "support_request_get",
         "Метод уточняется у заказчика. Получить заявку в службу поддержки по ЛС; заполнить слоты; в final_answer вернуть статус заявки."),
        ("Ошибки в сервисах",
         "Нету сейчас поля Email в интеграции",
         "", "Получить e-mail клиента для направления данных — метод уточняется у заказчика (нет в текущей модели ответа)",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "get_client_email",
         "Метод уточняется у заказчика (нет поля Email в текущей модели ответа). Получить e-mail клиента из ЛС; заполнить слот email."),
        ("Обновление ЛК",
         "Оплата поступила на ЛС, но не отобразилась в ЛК — проверка начислений/платежа",
         "", "Получить последние платежи по ЛС",
         "REST GET /{user_id}/transactions",
         "transactions[].transactions[]{payment; date}",
         "last_payment",
         "get_transactions",
         "Выполнить GET /transactions/{user_id}; получить последние платежи; заполнить слот last_payment; в final_answer объяснить сроки зачисления (до 3 рабочих дней)."),
        ("Акт сверки",
         "Оформление акта сверки взаиморасчетов",
         "", "Получить данные для акта сверки и оформить/отправить акт",
         "REST GET /{user_id}/reconciliation_act_data + POST /{user_id}/reconciliation_act",
         "contract_id; begin_date_default; end_date_default; items[]{name; value}",
         "—",
         "reconciliation_act",
         "Выполнить GET /reconciliation_act_data/{user_id}, затем POST /reconciliation_act/{user_id} для отправки акта сверки; заполнить слоты begin/end date, элементы селектора; в final_answer вернуть статус отправки."),
        ("Электронная квитанция",
         "Запрос на подключение квитанции по ЛС",
         "", "Подключение электронной квитанции по ЛС — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "billing_receipt_mode",
         "Метод уточняется у заказчика. Подключить электронную квитанцию по ЛС; в final_answer вернуть статус подключения."),
    ],

    "Скрипт ЮРИДИЧЕСКИЕ ЛИЦА": [
        ("уточнение № договора",
         "Спросить ИНН у человека либо получить по номеру телефона в лицевых счетах и запросить информацию о договорах этого ИНН",
         "", "Поиск договоров ЮЛ: по телефону — SOAP GetContractsInfo_By_Phone; поиск по ИНН — метод уточняется у заказчика",
         "SOAP IVR.asmx GetContractsInfo_By_Phone (+ поиск по ИНН: уточняется)",
         "ContractInfo{ID; No; Adress; AbonentName}; номера договоров по ИНН (уточняется)",
         "contract_no; user_id",
         "legal_contracts_by_inn",
         "Поиск договоров ЮЛ: по телефону — SOAP GetContractsInfo_By_Phone; поиск по ИНН — метод уточняется у заказчика. Заполнить слоты contract_no, user_id; в final_answer перечислить номера договоров."),
        ("уточнение № договора",
         "Запросить ИНН и проверить запросом наличие договора",
         "", "Проверка наличия договора по ИНН — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "legal_contract_check_inn",
         "Метод уточняется у заказчика. Проверить наличие договора по ИНН; заполнить слот contract_no; в final_answer вернуть найден/не найден."),
        ("уточнение № договора",
         "Инструкция по поиску в АСУСЭиРП",
         "", "Поиск договора в АСУСЭиРП (внешняя система) — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "asus_esrp_search",
         "Метод уточняется у заказчика. Поиск договора/ЛС ЮЛ в АСУСЭиРП; заполнить слоты; в final_answer вернуть результат поиска."),
        ("уточнение задолженности",
         "Уточнение задолженности по нежилому помещению",
         "", "Получить баланс по договору/телефону с разбивкой по поставщикам и услугам",
         "SOAP IVR.asmx GetContractsBalance_ByPhoneNumber / GetContractsBalance_ByContractID",
         "ContractNumber; ContragentName; DicServiceGroupName; Balance",
         "balance",
         "legal_balance",
         "SOAP GetContractsBalance_ByPhoneNumber/ByContractID; из ответа заполнить слот balance и перечень (ContractNumber, ContragentName, DicServiceGroupName, Balance); в final_answer озвучить сумму задолженности."),
        ("автодозвон",
         "По базе автодозвона найти клиента, озвучить номер договора, на который поступал автодозвон",
         "", "База автодозвона — метод уточняется у заказчика (вне REST API)",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "autodial_search",
         "Метод уточняется у заказчика. Поиск клиента в базе автодозвона; озвучить номер договора, на который поступал автодозвон; заполнить слот contract_no."),
        ("телефон расчетчика",
         "Найти ответственного расчетчика по договору",
         "", "Поиск расчетчика — АСУСЭиРП, метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "accountant_finder",
         "Метод уточняется у заказчика (АСУСЭиРП). Найти ответственного расчетчика по договору; заполнить слот accountant; в final_answer — ФИО/телефон расчетчика."),
        ("опломбировка",
         "Оформление заявки на опломбирование прибора учета (нежилое помещение)",
         "", "Создание заявки на опломбировку — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "specialist_request_create",
         "Метод уточняется у заказчика. Оформить заявку на опломбирование прибора учета (нежилое); в final_answer вернуть статус; id заявки в слот."),
        ("Передача показаний",
         "Передача показаний по электроэнергии/тепловой энергии (интегральный/интервальный ПУ)",
         "", "Передача показаний ЮЛ выполняется через ЛК для бизнеса, форму ВНЭ-1, ПО «Архивный считыватель» — вне API; метод уточняется у заказчика",
         "вне API (ЛК для бизнеса / ВНЭ-1 / Архивный считыватель)",
         "—",
         "—",
         "legal_readings_transfer",
         "Метод уточняется у заказчика. Передача показаний ЮЛ выполняется через ЛК для бизнеса, форму ВНЭ-1, ПО «Архивный считыватель»; агент — инструкция/перенаправление клиента."),
        ("личный кабинет",
         "Номер договора и ИНН ЮЛ в АСУСЭиРП",
         "", "Справка по ЛС ЮЛ из АСУСЭиРП — метод уточняется у заказчика",
         "метод уточняется у заказчика",
         "уточняется",
         "—",
         "asus_esrp_search",
         "Метод уточняется у заказчика. Получить номер договора и ИНН ЮЛ из АСУСЭиРП; заполнить слоты contract_no, inn; в final_answer озвучить."),
    ],
}

METHODS_ROWS = [
    ("info", "REST", "GET /api/service/sigurd/fl/{user_id}/info", "user_id",
     "Полная карточка ЛС: статус, баланс, собственник, услуги на расчетах, приборы учета, дом",
     "account{number; status; owner; balance; is_owner_registered; devices[]; accounts[]}; house{address}"),
    ("contracts_by_phone", "REST", "GET /api/service/sigurd/fl/contracts/by_phone/{phone}", "phone",
     "Поиск ЛС по номеру телефона",
     "UserId; UserPhone; WebProperties{IsRegistered; ContractNo; Address}"),
    ("contracts_by_number", "REST", "GET /api/service/sigurd/fl/contracts/by_number/{contractNo}", "contractNo",
     "Поиск ЛС по номеру лицевого счета",
     "UserId; WebProperties{ContractNo; Address}"),
    ("contracts_by_address_lastname", "REST", "GET /api/service/sigurd/fl/contracts/by_address_lastname",
     "lastName; cityName; streetName; houseName; flatName",
     "Поиск ЛС по адресу и фамилии владельца",
     "UserId; WebProperties{ContractNo}"),
    ("contracts_by_number_lastname", "REST", "GET /api/service/sigurd/fl/contracts/by_number_lastname/{contractNo}/{lastName}",
     "contractNo; lastName",
     "Поиск ЛС по номеру и фамилии владельца",
     "UserId; WebProperties{ContractNo}"),
    ("contracts_by_address", "REST", "GET /api/service/sigurd/fl/contracts/by_address",
     "cityName; streetName; houseName; search_all_flats",
     "Получение всех ЛС по адресу",
     "UserId[]; WebProperties{ContractNo; Address}[]"),
    ("disconnections_electro", "REST", "GET /api/service/sigurd/fl/{user_id}/disconnections_electro", "user_id",
     "Информация об отключении/ограничении электроэнергии и задолженности",
     "house_type; disconnection{text; type; work_type; date_start; date_end; duration; date_start_dt; date_end_dt}; debt"),
    ("disconnections_heat", "REST", "GET /api/service/sigurd/fl/{user_id}/disconnections_heat", "user_id",
     "Информация об отключении/ограничении теплоснабжения и ГВС",
     "disconnection{text; type; work_type; date_start; date_end; date_start_dt; date_end_dt; duration}; debt"),
    ("transactions", "REST", "GET /api/service/sigurd/fl/{user_id}/transactions", "user_id",
     "Платежи и поступления по ЛС (последняя оплата, способ поступления)",
     "transactions[].transactions[]{payment; date}"),
    ("devices", "REST", "GET /api/service/sigurd/fl/{user_id}/devices", "user_id",
     "Приборы учета в ЛС и их статусы",
     "devices[]{service_name; is_smart; is_installed; status; accepts_readings; readings_accept_type; last_reading; next_check}"),
    ("device_history", "REST", "GET /api/service/sigurd/fl/{user_id}/devices/history", "user_id",
     "История передачи показаний по приборам учета",
     "devices[]{service_name; last_reading; history[]{value; date; status_transfer}}"),
    ("disconnection_report_info", "REST", "GET /api/service/sigurd/fl/{user_id}/disconnection_report_info", "user_id",
     "Проверка возможности оформления заявки в КСРТ и данные для заявки",
     "request_allowed; данные для заявки"),
    ("disconnection_report_info", "REST", "POST /api/service/sigurd/fl/{user_id}/disconnection_report_info", "user_id; тело заявки",
     "Оформление заявки в КСРТ (аварийная заявка по отсутствию электроэнергии)",
     "id; status"),
    ("cancel_ksrt", "REST", "POST /api/service/sigurd/fl/{user_id}/cancel_ksrt", "user_id",
     "Отмена заявки КСРТ (например после оплаты задолженности)",
     "id; status"),
    ("notifications", "REST", "GET /api/service/sigurd/fl/{user_id}/notifications", "user_id",
     "Уведомления/оповещения по ЛС",
     "title; text; date"),
    ("notifications_electro", "REST", "GET /api/service/sigurd/fl/{user_id}/notifications_electro", "user_id",
     "Уведомления об отключениях электроэнергии",
     "title; text; date"),
    ("network_organization_data", "REST", "GET /api/service/sigurd/fl/{user_id}/network_organization_data", "user_id",
     "Сетевая организация по ЛС",
     "Name; Address; Contacts"),
    ("specialist_requests", "REST", "GET /api/service/sigurd/fl/{user_id}/specialist_requests", "user_id",
     "Заявки на вызов специалиста (обследование/опломбировка)",
     "requests[]{id; request_date; title; status; message}"),
    ("ksrt_requests", "REST", "GET /api/service/sigurd/fl/{user_id}/ksrt_requests", "user_id",
     "Заявки КСРТ",
     "requests[]{id; request_date; title; status; message}"),
    ("reconciliation_act_data", "REST", "GET /api/service/sigurd/fl/{user_id}/reconciliation_act_data", "user_id",
     "Данные для оформления акта сверки",
     "contract_id; begin_date_min; begin_date_default; end_date_default; selector_label; count_peni_default; items[]{name; value}"),
    ("reconciliation_act", "REST", "POST /api/service/sigurd/fl/{user_id}/reconciliation_act", "user_id; тело запроса",
     "Оформление и отправка акта сверки",
     "id; status"),
    ("send_account_number", "REST", "POST /api/service/sigurd/fl/{user_id}/send_account_number", "user_id; тело запроса",
     "Отправка номера ЛС на email/SMS клиенту",
     "status"),
    ("FindAllByContractNumber", "SOAP", "IVR.asmx FindAllByContractNumber", "contractNumberDigits",
     "Поиск ЛС по цифровой части номера",
     "ContractInfo[]"),
    ("GetContractsInfo_By_Phone", "SOAP", "IVR.asmx GetContractsInfo_By_Phone", "PhoneNumber",
     "Поиск ЛС по номеру телефона",
     "ContractInfo{ID; No; Adress; City; Residents; FullArea; DateUpdate; Stove; Status; AbonentName; FirstName; Patronimic; DivisionID}"),
    ("GetContractsBalance_ByPhoneNumber", "SOAP", "IVR.asmx GetContractsBalance_ByPhoneNumber", "phoneNumber",
     "Баланс по ЛС абонента с разбивкой по поставщикам и услугам",
     "ContractAddressAndBalanceByDicServiceGroupAndContragent[]{ContractNumber; Settlement; Address; ContragentName; DicServiceGroupName; Balance}"),
    ("GetContractsBalance_ByContractID", "SOAP", "IVR.asmx GetContractsBalance_ByContractID", "ContractID",
     "Баланс по ЛС с разбивкой по поставщикам и услугам",
     "ContractAddressAndBalanceByDicServiceGroupAndContragent[]"),
    ("GetMDInfo_By_ContractIDAndNomenclatureCode", "SOAP", "IVR.asmx GetMDInfo_By_ContractIDAndNomenclatureCode",
     "ContractStrGUID; NomenclatureCode",
     "Информация о приборе учёта (серийный номер, шкалы, показания)",
     "данные прибора учета; шкалы с последними показаниями"),
    ("GetMDsWarningInfo_ByTel", "SOAP", "IVR.asmx GetMDsWarningInfo_ByTel", "telephoneNumber",
     "Предупреждения по приборам учета (по телефону)",
     "MDWarningInfo[]{AddressOfMD; DicServiceGroupNumber; LastReadings; NextVerificationDeadline; IsReadingsTooOld; VerificationStatus}"),
    ("GetMDsWarningInfo_ByContractID", "SOAP", "IVR.asmx GetMDsWarningInfo_ByContractID", "contractID",
     "Предупреждения по приборам учета (по ЛС)",
     "MDWarningInfo[]{AddressOfMD; DicServiceGroupNumber; LastReadings; NextVerificationDeadline; IsReadingsTooOld; VerificationStatus}"),
    ("InputOfReadingsWithDot", "SOAP", "IVR.asmx InputOfReadingsWithDot", "MDScale_ID; MDScale_NewReadings",
     "Передача показаний по шкале прибора учета",
     "результат передачи"),
]


def build():
    wb = Workbook()
    wb.remove(wb.active)
    init_methods_meta()

    items = collect_enriched()
    usage = compute_usage(items)

    ws_id = wb.create_sheet("Идентификация клиента")
    style_header(ws_id, HEADER_SCRIPT, WIDTHS_SCRIPT)
    write_rows(ws_id, "Идентификация клиента", IDENT_ROWS)

    add_methods_sheet(wb, METHODS_ROWS, usage)
    build_slots_sheet(wb, items)
    build_agents_sheet(wb, items)

    for title, rows in SCRIPT_ROWS.items():
        add_script_sheet(wb, title, rows)

    order = (["Идентификация клиента", "Методы API", "Слоты", "Агенты"]
             + list(SCRIPT_ROWS.keys()))
    wb._sheets = [wb[n] for n in order]

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print("OK:", OUT)
    print("Листы:", ", ".join(wb.sheetnames))
    total = sum(wb[n].max_row - 1 for n in order)
    print("Всего строк:", total)


if __name__ == "__main__":
    build()
